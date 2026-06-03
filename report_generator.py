# -*- coding: utf-8 -*-
"""
Генерация Excel-отчёта по результатам автоматических проверок.

Используется openpyxl: лист «Общая сводка» и лист «Дефекты».
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class DefectRecord:
    """Одна строка для листа «Дефекты»."""

    defect_id: str
    defect_type: str
    priority: str
    description: str
    location: str
    screenshot_path: str
    recommendation: str


@dataclass
class RunSummary:
    """Сводка прогона для листа «Общая сводка»."""

    started_at: datetime
    finished_at: datetime
    base_url: str
    total_checks: int
    passed: int
    failed: int
    critical_issues: List[str] = field(default_factory=list)
    notes: List[str] = field(default_factory=list)


def _header_fill():
    return PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")


def _header_font():
    return Font(color="FFFFFF", bold=True)


def write_excel_report(
    output_path: Path,
    summary: RunSummary,
    defects: List[DefectRecord],
) -> None:
    """
    Создаёт файл .xlsx с двумя листами: «Общая сводка» и «Дефекты».

    :param output_path: путь к report_YYYYMMDD_HHMMSS.xlsx
    :param summary: агрегированные метрики прогона
    :param defects: список найденных дефектов
    """
    wb = Workbook()

    # --- Лист 1: Общая сводка ---
    ws1 = wb.active
    ws1.title = "Общая сводка"

    duration_sec = (summary.finished_at - summary.started_at).total_seconds()
    rows = [
        ("Параметр", "Значение"),
        ("URL", summary.base_url),
        ("Начало", summary.started_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Окончание", summary.finished_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Длительность (сек)", round(duration_sec, 2)),
        ("Всего проверок (условных единиц)", summary.total_checks),
        ("Пройдено", summary.passed),
        ("Не пройдено", summary.failed),
        ("Критических проблем (кол-во)", len(summary.critical_issues)),
    ]

    for r, (label, value) in enumerate(rows, start=1):
        ws1.cell(row=r, column=1, value=label)
        ws1.cell(row=r, column=2, value=value)
        if r == 1:
            for c in (1, 2):
                cell = ws1.cell(row=r, column=c)
                cell.fill = _header_fill()
                cell.font = _header_font()
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws1.cell(row=len(rows) + 2, column=1, value="Критические проблемы (список):")
    ws1.cell(row=len(rows) + 2, column=1).font = Font(bold=True)

    start_crit = len(rows) + 3
    if summary.critical_issues:
        for i, text in enumerate(summary.critical_issues, start=0):
            ws1.cell(row=start_crit + i, column=1, value=text)
            ws1.merge_cells(
                start_row=start_crit + i,
                start_column=1,
                end_row=start_crit + i,
                end_column=3,
            )
    else:
        ws1.cell(row=start_crit, column=1, value="— нет —")

    note_row = start_crit + max(len(summary.critical_issues), 1) + 2
    ws1.cell(row=note_row, column=1, value="Примечания (не дефекты):")
    ws1.cell(row=note_row, column=1).font = Font(bold=True)
    if summary.notes:
        for i, text in enumerate(summary.notes, start=0):
            ws1.cell(row=note_row + 1 + i, column=1, value=text)
            ws1.merge_cells(
                start_row=note_row + 1 + i,
                start_column=1,
                end_row=note_row + 1 + i,
                end_column=3,
            )
    else:
        ws1.cell(row=note_row + 1, column=1, value="—")

    ws1.column_dimensions["A"].width = 36
    ws1.column_dimensions["B"].width = 60

    # --- Лист 2: Дефекты ---
    ws2 = wb.create_sheet("Дефекты")
    defect_headers = [
        "ID",
        "Тип",
        "Приоритет",
        "Описание",
        "Локация (URL / селектор)",
        "Скриншот (путь)",
        "Рекомендация по исправлению",
    ]
    for col, h in enumerate(defect_headers, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row_idx, d in enumerate(defects, start=2):
        ws2.cell(row=row_idx, column=1, value=d.defect_id)
        ws2.cell(row=row_idx, column=2, value=d.defect_type)
        ws2.cell(row=row_idx, column=3, value=d.priority)
        ws2.cell(row=row_idx, column=4, value=d.description)
        ws2.cell(row=row_idx, column=5, value=d.location)
        ws2.cell(row=row_idx, column=6, value=d.screenshot_path)
        ws2.cell(row=row_idx, column=7, value=d.recommendation)
        for c in range(1, 8):
            ws2.cell(row=row_idx, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    widths = [10, 18, 12, 50, 40, 35, 40]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
